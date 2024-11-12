from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from django.http import Http404, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.utils.timezone import make_aware
from django.views.generic import View

from CashRegister.models import CashRegister
from Reports.forms import CashRegisterReportForm, SalesReportForm, ProductsReportForm
from Sales.models import Sales
from Products.models import Products
from utils.user_utils import validate_user_acess

from openpyxl import Workbook
import io


@method_decorator(
    login_required(login_url='users:login', redirect_field_name='next'),
    name='dispatch'
)
class SelectReportClassView(View):
    def get(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        title = 'Selecionar'
        subtitle = 'Relatório'

        return render(
            self.request,
            'reports/pages/select_report.html',
            context={
                'site_title': f'{title} {subtitle}',
                'page_title': title,
                'page_subtitle': subtitle,
            }
        )


@method_decorator(
    login_required(login_url='users:login', redirect_field_name='next'),
    name='dispatch'
)
class CashRegisterReport(View):
    def render_page(self, form: CashRegisterReportForm):
        title = 'Relatório'
        subtitle = 'de Caixa'

        return render(
            self.request,
            'reports/pages/cashregister_report.html',
            context={
                'form': form,
                'site_title': f'{title} {subtitle}',
                'page_title': title,
                'page_subtitle': subtitle,
                'is_report_page': True,
            }
        )

    def get(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        return self.render_page(CashRegisterReportForm())

    def post(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        form = CashRegisterReportForm(
            data=self.request.POST or None,
            files=self.request.FILES or None
        )

        if form.is_valid():
            date_initial = form.cleaned_data.get('date_initial')
            date_final = form.cleaned_data.get('date_final')

            if date_initial:
                date_initial = datetime.combine(
                    date_initial, datetime.min.time()
                )
                date_initial = make_aware(date_initial)

            if date_final:
                date_final = datetime.combine(date_final, datetime.max.time())
                date_final = make_aware(date_final)

            cashregisters = CashRegister.objects.filter(
                is_open=False,
                open_date__range=(date_initial, date_final)
            ).prefetch_related(Prefetch(
                'sales', queryset=Sales.objects.select_related('payment_type')
            ),)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Relatório de Caixa"

            worksheet.append(["ID", "Data abertura", "Data fechamento", "Saldo", "QTD Vendas"])

            for cashregister in cashregisters:
                open_date = cashregister.open_date.replace(tzinfo=None) if cashregister.open_date is not None else cashregister.open_date
                close_date = cashregister.close_date.replace(tzinfo=None) if cashregister.close_date is not None else cashregister.close_date
                sales = cashregister.sales.all()
                sales_count = len(sales) if sales else 0

                worksheet.append([
                    cashregister.id,
                    open_date,
                    close_date,
                    cashregister.cash,
                    sales_count
                ])

            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Relatorio_de_Caixa.xlsx'
            return response

        return self.render_page(form)


class SalesReport(View):
    def render_page(self, form: SalesReportForm):
        title = 'Relatório'
        subtitle = 'de Vendas'

        return render(
            self.request,
            'reports/pages/sales_report.html',
            context={
                'form': form,
                'site_title': f'{title} {subtitle}',
                'page_title': title,
                'page_subtitle': subtitle,
                'is_report_page': True,
            }
        )

    def get(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        return self.render_page(SalesReportForm())

    def post(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        form = SalesReportForm(
            data=self.request.POST or None,
            files=self.request.FILES or None
        )

        if form.is_valid():
            date_initial = form.cleaned_data.get('date_initial')
            date_final = form.cleaned_data.get('date_final')

            if date_initial:
                date_initial = datetime.combine(
                    date_initial, datetime.min.time()
                )
                date_initial = make_aware(date_initial)

            if date_final:
                date_final = datetime.combine(date_final, datetime.max.time())
                date_final = make_aware(date_final)

            sales = Sales.objects.filter(
                created_at__range=(date_initial, date_final)
            ).select_related('payment_type', 'schedule').prefetch_related(
                Prefetch('products'), Prefetch('schedule__services')
            )

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Relatório de Vendas"

            worksheet.append(["ID Venda", "Data", "Cliente", "Tipo de Venda", "Itens", "Total (R$)", "Tipo de Pagamento"])

            for sale in sales:
                sale_type = "Serviço" if sale.schedule else "Produto"

                if sale_type == "Produto":
                    items = ", ".join([product.product_name for product in sale.products.all()])
                else:
                    items = ", ".join([service.service_name for service in sale.schedule.services.all()])

                worksheet.append([
                    sale.id,
                    sale.created_at.replace(tzinfo=None),
                    sale.user.first_name,
                    sale_type,
                    items,
                    sale.total_price,
                    sale.payment_type.payment_name
                ])

            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Relatorio_de_Vendas.xlsx'
            return response

        return self.render_page(form)
    

class ProductsReport(View):
    def render_page(self, form: ProductsReportForm):
        title = 'Relatório'
        subtitle = 'de Produtos'

        return render(
            self.request,
            'reports/pages/products_report.html',
            context={
                'form': form,
                'site_title': f'{title} {subtitle}',
                'page_title': title,
                'page_subtitle': subtitle,
                'is_report_page': True,
            }
        )

    def get(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        return self.render_page(ProductsReportForm())

    def post(self, *args, **kwargs):
        if not validate_user_acess(self.request):
            return redirect(reverse('users:no-permission'))

        form = ProductsReportForm(
            data=self.request.POST or None,
            files=self.request.FILES or None
        )

        if form.is_valid():
            is_active = form.cleaned_data.get('is_active')
            show_in_showcase = form.cleaned_data.get('show_in_showcase')

            products = Products.objects.all()

            if is_active != 'Todos':
                products = products.filter(is_active=is_active == 'True')
            if show_in_showcase != 'Todos':
                products = products.filter(show=show_in_showcase == 'True')

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Relatório de Produtos"

            worksheet.append(["ID Produto", "Nome", "Categoria", "Preço (R$)", "Status", "Mostrar na Vitrine", "Data de Criação"])

            for product in products:
                status = "Ativo" if product.is_active else "Inativo"
                showcase = "Sim" if product.show else "Não"
                worksheet.append([
                    product.id,
                    product.product_name,
                    product.product_category.category_name if product.product_category else "Sem Categoria",
                    product.price,
                    status,
                    showcase,
                    product.created_at.strftime("%d/%m/%Y"),
                ])

            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=Relatorio_de_Produtos.xlsx'
            return response

        return self.render_page(form)